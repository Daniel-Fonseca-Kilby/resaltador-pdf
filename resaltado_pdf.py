"""
resaltado_pdf.py

Disclaimer: buscar texto (nombres) dentro de un PDF y agregarle
anotaciones de resaltado ("highlight"), SIN modificar el resto del documento
(mismo encabezado, mismo formato, mismas fuentes, etc.).

"""

import io
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pymupdf as fitz  # PyMuPDF (alias 'fitz' por compatibilidad con ejemplos/documentación)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass
class ResultadoResaltado:
    """Mensaje de resultado que devuele el aplicativo al usuario"""
    archivo: str
    coincidencias_por_nombre: dict  # {"JUAN PEREZ": 3, "MARIA LOPEZ": 0}
    ruta_salida: str


def _normalizar(texto: str) -> str:
    """Normaliza el texto para tolerar diferencias de mayúsculas/minúsculas y tildes. Esto ayuda a"""
    texto = texto.strip().upper()
    texto_sin_tildes = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto_sin_tildes if not unicodedata.combining(c))


def _variantes_ene(texto: str) -> list[str]:
    """La Ñ/ñ a veces sale como espacio en blanco en el texto del PDF (pasa con MNK)."""
    if "Ñ" not in texto and "ñ" not in texto:
        return []
    return [
        texto.replace("Ñ", " ").replace("ñ", " "),
        texto.replace("Ñ", "").replace("ñ", ""),
    ]


def _buscar_con_variantes(pagina, texto: str, textpage=None):
    """Prueba la forma exacta, sin tildes, y variantes de Ñ si aplica.
    'textpage' (opcional) evita que PyMuPDF reextraiga el texto de la
    página en cada búsqueda -se puede sacar una vez con page.get_textpage()
    y reusarla para todas las búsquedas de esa página."""
    coincidencias = pagina.search_for(texto, quads=False, textpage=textpage)
    if coincidencias:
        return coincidencias

    coincidencias = pagina.search_for(_normalizar(texto), quads=False, textpage=textpage)
    if coincidencias:
        return coincidencias

    for variante in _variantes_ene(texto):
        coincidencias = pagina.search_for(variante, quads=False, textpage=textpage)
        if coincidencias:
            return coincidencias

    return []


_TOLERANCIA_FILA = 6  # puntos: cuánto pueden variar en Y las palabras de una misma fila


def _buscar_por_fila(pagina, palabras: list[str], textpage=None):
    """Busca cada palabra suelta y agrupa las que caen en la misma fila
    (nombre y apellido en columnas separadas, por ejemplo)."""
    rects_por_palabra = []
    for palabra in palabras:
        rects = _buscar_con_variantes(pagina, palabra, textpage=textpage)
        if not rects:
            return []
        rects_por_palabra.append(rects)

    indice_ancla = min(range(len(rects_por_palabra)), key=lambda i: len(rects_por_palabra[i]))
    filas_encontradas = []
    for rect_ancla in rects_por_palabra[indice_ancla]:
        fila = [rect_ancla]
        completa = True
        for i, rects in enumerate(rects_por_palabra):
            if i == indice_ancla:
                continue
            candidato = next((r for r in rects if abs(r.y0 - rect_ancla.y0) <= _TOLERANCIA_FILA), None)
            if candidato is None:
                completa = False
                break
            fila.append(candidato)
        if completa:
            filas_encontradas.append(fila)
    return filas_encontradas


def resaltar_nombres_en_pdf(ruta_pdf: str, nombres: list[str], ruta_salida: str) -> ResultadoResaltado:
    """Resalta cada nombre de la lista en el PDF y guarda la copia en ruta_salida."""
    documento = fitz.open(ruta_pdf)
    if documento.is_encrypted:
        documento.close()
        raise ValueError(
            "El PDF está protegido con contraseña. Quite la protección e inténtelo de nuevo."
        )

    conteo = {nombre: 0 for nombre in nombres}

    for pagina in documento:
        textpage = pagina.get_textpage()  # se extrae 1 vez y se reusa en todas las búsquedas de la página
        for nombre in nombres:
            nombre_limpio = nombre.strip()
            if not nombre_limpio:
                continue

            coincidencias = _buscar_con_variantes(pagina, nombre_limpio, textpage=textpage)

            if coincidencias:
                for rect in coincidencias:
                    anotacion = pagina.add_highlight_annot(rect)
                    anotacion.update()
                    conteo[nombre] += 1
                continue

            # nombre completo no aparece junto -> capaz nombre/apellido
            # quedaron en columnas distintas, se busca palabra por palabra
            palabras = [p for p in nombre_limpio.split() if len(p) >= 3]
            for fila in _buscar_por_fila(pagina, palabras, textpage=textpage):
                for rect in fila:
                    anotacion = pagina.add_highlight_annot(rect)
                    anotacion.update()
                conteo[nombre] += 1

    documento.save(ruta_salida)
    documento.close()

    return ResultadoResaltado(
        archivo=Path(ruta_pdf).name,
        coincidencias_por_nombre=conteo,
        ruta_salida=ruta_salida,
    )


def _nombre_archivo_seguro(texto: str) -> str:
    limpio = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in texto).strip()
    return limpio or "SIN_CLIENTE"


def _normalizar_cedula(digitos: str) -> str:
    """Quita ceros a la izquierda: una celda numérica del Excel (en vez de
    texto) puede perder el cero inicial de la cédula, y con esto igual calza
    contra la que viene completa desde el PDF."""
    return digitos.lstrip("0") or "0"


def _coincide_cliente(digitos_palabra: str, mapa_cedulas: dict) -> list[str]:
    """Devuelve todos los clientes a los que pertenece esta cédula (relación
    1 a N: un oficial puede cubrir turnos en más de un cliente durante la
    misma quincena). Tolera que la planilla anteponga un dígito de tipo de
    id (ej. CCSS: '0-303370238') y que a la cédula le falten ceros iniciales
    frente a la del Excel. 'mapa_cedulas' debe venir con las claves ya
    normalizadas (ver _normalizar_cedula)."""
    clientes = mapa_cedulas.get(_normalizar_cedula(digitos_palabra))
    if not clientes and len(digitos_palabra) > 1:
        clientes = mapa_cedulas.get(_normalizar_cedula(digitos_palabra[1:]))
    return clientes or []


# frases que solo salen en la fila de títulos de cada formato conocido
_PERFILES_ENCABEZADO = {
    "ccss": ["APELLIDOS Y NOMBRES", "OBSERVACIONES"],
    "mnk": ["IDENTIFICACIÓN", "OBSERVACIÓN"],
    "ins": ["Salario Reportado", "Descripción Ocupación"],
}


def _techo_por_anclas(pagina, anclas: list[str], textpage=None) -> float | None:
    """Y justo debajo de la fila donde aparecen las anclas, con margen."""
    y1_maximo = None
    for ancla in anclas:
        coincidencias = pagina.search_for(ancla, quads=False, textpage=textpage)
        if not coincidencias:
            coincidencias = pagina.search_for(_normalizar(ancla), quads=False, textpage=textpage)
        for rect in coincidencias:
            if y1_maximo is None or rect.y1 > y1_maximo:
                y1_maximo = rect.y1
    if y1_maximo is None:
        return None
    return y1_maximo + 14


def _techo_de_datos(pagina, formato: str = "auto", textpage=None) -> float | None:
    """Y donde empieza la tabla de empleados (debajo de empresa/período/títulos).
    Prueba primero las anclas del formato indicado y despues las de los
    demás perfiles conocidos por si el desplegable quedó mal puesto -cada
    proveedor tiene frases únicas asi que no calzan cruzadas. Si nada
    calza cae al detector de tablas de PyMuPDF, y si tampoco encuentra
    nada mejor no mostrar encabezado que arriesgarse a mostrar una fila
    de otra persona."""
    perfiles_a_probar = []
    anclas_formato = _PERFILES_ENCABEZADO.get(formato)
    if anclas_formato:
        perfiles_a_probar.append(anclas_formato)
    for anclas_perfil in _PERFILES_ENCABEZADO.values():
        if anclas_perfil not in perfiles_a_probar:
            perfiles_a_probar.append(anclas_perfil)

    for anclas in perfiles_a_probar:
        techo = _techo_por_anclas(pagina, anclas, textpage=textpage)
        if techo is not None:
            return techo

    for estrategia in ("lines_strict", "lines", "text"):
        try:
            tablas = pagina.find_tables(strategy=estrategia).tables
        except Exception:
            continue
        if not tablas:
            continue
        tabla_datos = max(tablas, key=lambda t: len(t.rows))
        if tabla_datos.rows:
            return fitz.Rect(tabla_datos.rows[0].bbox).y0
    return None


_MARGEN_PAGINA = 24  # puntos de margen arriba/abajo en cada página de salida
_ESPACIO_ENTRE_FILAS = 3  # separación vertical entre filas apiladas


def resaltar_por_cedula_y_exportar_por_cliente(
    rutas_pdfs: list[str],
    registros: list[dict],
    carpeta_salida: str,
    formato: str = "auto",
    resaltar_filas: bool = True,
) -> dict:
    """Busca las cédulas de 'registros' en los PDFs, recorta cada fila
    encontrada y arma un PDF por cliente, apilando filas sin dejar huecos.
    'formato' solo ayuda a ubicar el encabezado de cada página (ver
    _PERFILES_ENCABEZADO / _techo_de_datos) para no arrastrar filas de
    gente que no está en el Excel. 'resaltar_filas' controla si cada fila
    se marca con la franja amarilla o se deja tal cual salió de la
    planilla original -algunos clientes (auditorías, entidades públicas)
    piden el documento limpio, sin marcas encima.

    Una misma cédula puede estar asignada a más de un cliente en 'registros'
    (relación 1 a N -un oficial que cubrió turnos en varios puestos durante
    la quincena): en ese caso su fila se agrega al PDF de cada cliente al
    que está asignado, no solo al último que aparece en el Excel.

    El encabezado se repite una vez por archivo/póliza (no por página de
    origen): si un cliente tiene oficiales en varias páginas del mismo PDF
    solo se lleva el encabezado una vez, pero si cambia de archivo (otra
    póliza) se fuerza una hoja de salida limpia con su propio encabezado.
    Y si esa póliza termina desbordando a una segunda hoja de salida, el
    encabezado se repite ahí también para no perder el contexto.

    Ese encabezado SIEMPRE se recorta de la página 1 del archivo, nunca de
    la página donde arranca cada cliente: en reportes de varias páginas
    (ej. MNK) solo la primera trae el logo/título/fecha completos, las
    siguientes repiten nada más la fila de títulos de columna -si se usara
    la página del cliente, uno que empieza en la página 2 se llevaría el
    encabezado recortado.

    Segunda pasada, solo por nombre: los registros que no calzaron por
    cédula se vuelven a intentar buscando su nombre completo (rescata,
    ej., cuando el Excel trae un DIMEX pero la planilla imprime el número
    de CCSS de esa persona). Solo se acepta si el nombre aparece en
    exactamente una fila de todo el lote y esa fila no es ya de otro
    empleado conocido -si no, se deja como no encontrado."""
    # una entrada por cada par (cédula, cliente) único del Excel -si la
    # misma fila viene duplicada se conserva solo la primera aparición
    registros_unicos: dict[tuple[str, str], dict] = {}
    for r in registros:
        cedula, cliente = r.get("cedula"), r.get("cliente")
        if not cedula or not cliente:
            continue
        clave = (_normalizar_cedula(cedula), cliente)
        if clave not in registros_unicos:
            registros_unicos[clave] = {"cedula": cedula, "cliente": cliente, "nombre": r.get("nombre", "")}

    mapa_cedulas: dict[str, list[str]] = {}
    for clave_cedula, cliente in registros_unicos:
        clientes = mapa_cedulas.setdefault(clave_cedula, [])
        if cliente not in clientes:
            clientes.append(cliente)

    estado_por_cliente: dict[str, dict] = {}
    # (clave_cedula, cliente) -> set de nombres de archivo/póliza donde se encontró
    polizas_encontradas: dict[tuple[str, str], set] = {}
    errores_por_archivo: dict[str, str] = {}

    def _obtener_estado(cliente: str, pagina_origen) -> dict:
        estado = estado_por_cliente.get(cliente)
        if estado is None:
            estado = {
                "documento": fitz.open(),
                "pagina": None,
                "y": 0.0,
                "ancho": pagina_origen.rect.width,
                "alto": pagina_origen.rect.height,
                "archivo_actual": None,  # ruta_pdf de la póliza que se está volcando ahora
                "encabezado_actual": None,  # bloque a repetir si la póliza desborda una hoja
            }
            estado_por_cliente[cliente] = estado
        return estado

    def _agregar_bloque(
        estado: dict,
        documento_origen,
        pagina_origen,
        franja: "fitz.Rect",
        resaltar: bool,
        repetir_encabezado: bool = True,
    ) -> None:
        escala = estado["ancho"] / pagina_origen.rect.width
        alto_bloque = franja.height * escala
        if alto_bloque <= 0:
            return

        if estado["pagina"] is None or estado["y"] + alto_bloque > estado["alto"] - _MARGEN_PAGINA:
            estado["pagina"] = estado["documento"].new_page(width=estado["ancho"], height=estado["alto"])
            estado["y"] = _MARGEN_PAGINA
            encabezado = estado["encabezado_actual"]
            if repetir_encabezado and encabezado is not None:
                # la póliza sigue en la hoja de salida siguiente -se repite el encabezado
                _agregar_bloque(
                    estado, encabezado["documento"], encabezado["pagina"], encabezado["franja"],
                    resaltar=False, repetir_encabezado=False,
                )

        destino = fitz.Rect(0, estado["y"], estado["ancho"], estado["y"] + alto_bloque)
        estado["pagina"].show_pdf_page(destino, documento_origen, pagina_origen.number, clip=franja)
        if resaltar:
            anotacion = estado["pagina"].add_highlight_annot(destino)
            anotacion.update()
        estado["y"] += alto_bloque + _ESPACIO_ENTRE_FILAS

    for ruta_pdf in rutas_pdfs:
        nombre_archivo = Path(ruta_pdf).name
        try:
            documento = fitz.open(ruta_pdf)
        except Exception as error:
            errores_por_archivo[nombre_archivo] = f"No se pudo abrir el archivo: {error}"
            continue

        if documento.is_encrypted:
            documento.close()
            errores_por_archivo[nombre_archivo] = (
                "El PDF está protegido con contraseña. Quite la protección e inténtelo de nuevo."
            )
            continue

        # encabezado de la póliza: SIEMPRE se saca de la página 1 del archivo,
        # no de la página donde arranca cada cliente -en reportes de varias
        # páginas (ej. MNK) solo la primera trae el logo/título/fecha
        # completos, las siguientes repiten nada más la fila de títulos de
        # columna. Se calcula una sola vez por archivo (no por cliente).
        encabezado_archivo = None
        encabezado_calculado = False

        try:
            for pagina in documento:
                textpage = pagina.get_textpage()
                palabras = pagina.get_text("words", textpage=textpage)

                franjas_vistas: dict[str, list] = {}
                for x0, y0, x1, y1, palabra, *_resto in palabras:
                    digitos = "".join(c for c in palabra if c.isdigit())
                    if not digitos:
                        continue
                    clientes = _coincide_cliente(digitos, mapa_cedulas)
                    if not clientes:
                        continue
                    clave_cedula = _normalizar_cedula(digitos)
                    if clave_cedula not in mapa_cedulas and len(digitos) > 1:
                        clave_cedula = _normalizar_cedula(digitos[1:])

                    fila_y0 = min(w[1] for w in palabras if abs(w[1] - y0) <= _TOLERANCIA_FILA)
                    fila_y1 = max(w[3] for w in palabras if abs(w[1] - y0) <= _TOLERANCIA_FILA)
                    franja = fitz.Rect(pagina.rect.x0, fila_y0 - 2, pagina.rect.x1, fila_y1 + 2)

                    # un mismo oficial puede estar asignado a varios clientes
                    # (relación 1 a N) -su fila se agrega al PDF de cada uno
                    for cliente in clientes:
                        polizas_encontradas.setdefault((clave_cedula, cliente), set()).add(nombre_archivo)

                        vistas = franjas_vistas.setdefault(cliente, [])
                        if franja in vistas:
                            continue
                        vistas.append(franja)

                        estado = _obtener_estado(cliente, pagina)

                        if estado["archivo_actual"] != ruta_pdf:
                            # nueva póliza para este cliente -hoja limpia y encabezado propio
                            # (si el cliente ya venía de otro archivo se repite este mismo
                            # bloque de encabezado en cada hoja que la póliza necesite)
                            if not encabezado_calculado:
                                primera_pagina = documento[0]
                                techo_pagina1 = _techo_de_datos(primera_pagina, formato)
                                if techo_pagina1 is not None and techo_pagina1 > 4:
                                    encabezado_archivo = {
                                        "documento": documento,
                                        "pagina": primera_pagina,
                                        "franja": fitz.Rect(
                                            primera_pagina.rect.x0, 0, primera_pagina.rect.x1, techo_pagina1 - 2
                                        ),
                                    }
                                encabezado_calculado = True
                            estado["pagina"] = None
                            estado["archivo_actual"] = ruta_pdf
                            estado["encabezado_actual"] = encabezado_archivo
                            if encabezado_archivo is not None:
                                _agregar_bloque(
                                    estado, encabezado_archivo["documento"], encabezado_archivo["pagina"],
                                    encabezado_archivo["franja"], resaltar=False, repetir_encabezado=False,
                                )

                        _agregar_bloque(estado, documento, pagina, franja, resaltar=resaltar_filas)
        except Exception as error:
            errores_por_archivo[nombre_archivo] = f"No se pudo procesar el archivo: {error}"
        finally:
            documento.close()

    # segunda pasada, solo por nombre: para los que no calzaron por cédula
    # en la primera pasada (ej. el Excel trae un DIMEX pero la planilla
    # imprime el número de CCSS de esa persona). Es más arriesgada que
    # buscar por cédula -nombres repetidos, coincidencias parciales-, así
    # que se aplica con baranda: solo cuenta si el nombre completo aparece
    # en EXACTAMENTE una fila de todo el lote, y esa fila no tiene ya una
    # cédula que pertenezca a otro empleado del Excel (no le "roba" la fila
    # a otra persona que solo comparte nombre).
    encontrados_por_nombre: set = set()
    pendientes_por_nombre: dict[tuple[str, str], list[str]] = {}
    for clave, datos in registros_unicos.items():
        if polizas_encontradas.get(clave):
            continue
        palabras_nombre = [p for p in _normalizar(datos.get("nombre") or "").split() if len(p) >= 3]
        if len(palabras_nombre) >= 2:
            pendientes_por_nombre[clave] = palabras_nombre

    if pendientes_por_nombre:
        candidatos_por_pendiente: dict[tuple[str, str], list[tuple[str, int, tuple]]] = {}
        for ruta_pdf in rutas_pdfs:
            try:
                documento = fitz.open(ruta_pdf)
            except Exception:
                continue
            if documento.is_encrypted:
                documento.close()
                continue
            try:
                for pagina in documento:
                    textpage = pagina.get_textpage()
                    palabras_pagina = pagina.get_text("words", textpage=textpage)
                    for clave, palabras_nombre in pendientes_por_nombre.items():
                        for fila_rects in _buscar_por_fila(pagina, palabras_nombre, textpage=textpage):
                            fila_y0 = min(r.y0 for r in fila_rects)
                            fila_y1 = max(r.y1 for r in fila_rects)

                            # ¿esta fila ya tiene su propia cédula, de otro
                            # empleado conocido del Excel? si es así, no es
                            # esta persona -aunque el nombre haya calzado
                            cedula_de_la_fila = None
                            for wx0, wy0, wx1, wy1, wpalabra, *_resto in palabras_pagina:
                                if abs(wy0 - fila_y0) > _TOLERANCIA_FILA:
                                    continue
                                wdigitos = "".join(c for c in wpalabra if c.isdigit())
                                if not wdigitos:
                                    continue
                                wclave_cedula = _normalizar_cedula(wdigitos)
                                if wclave_cedula not in mapa_cedulas and len(wdigitos) > 1:
                                    wclave_cedula = _normalizar_cedula(wdigitos[1:])
                                if wclave_cedula in mapa_cedulas:
                                    cedula_de_la_fila = wclave_cedula
                                    break
                            if cedula_de_la_fila is not None and cedula_de_la_fila != clave[0]:
                                continue

                            franja = fitz.Rect(pagina.rect.x0, fila_y0 - 2, pagina.rect.x1, fila_y1 + 2)
                            candidatos_por_pendiente.setdefault(clave, []).append(
                                (ruta_pdf, pagina.number, (franja.x0, franja.y0, franja.x1, franja.y1))
                            )
            except Exception:
                pass
            finally:
                documento.close()

        for clave, candidatos in candidatos_por_pendiente.items():
            if len(candidatos) != 1:
                continue  # ninguna fila, o ambiguo entre varias -no se arriesga
            ruta_pdf_ganador, numero_pagina, coords = candidatos[0]
            cliente = clave[1]
            franja = fitz.Rect(*coords)

            try:
                documento_ganador = fitz.open(ruta_pdf_ganador)
            except Exception:
                continue
            try:
                pagina_ganadora = documento_ganador[numero_pagina]
                estado = _obtener_estado(cliente, pagina_ganadora)

                primera_pagina_ganadora = documento_ganador[0]
                techo_ganador = _techo_de_datos(primera_pagina_ganadora, formato)
                encabezado_ganador = None
                if techo_ganador is not None and techo_ganador > 4:
                    encabezado_ganador = {
                        "documento": documento_ganador,
                        "pagina": primera_pagina_ganadora,
                        "franja": fitz.Rect(
                            primera_pagina_ganadora.rect.x0, 0, primera_pagina_ganadora.rect.x1, techo_ganador - 2
                        ),
                    }
                # se refresca siempre (con este documento, todavía abierto)
                # para no dejar una referencia muerta de un archivo ya cerrado
                estado["encabezado_actual"] = encabezado_ganador

                if estado["archivo_actual"] != ruta_pdf_ganador:
                    estado["pagina"] = None
                    estado["archivo_actual"] = ruta_pdf_ganador
                    if encabezado_ganador is not None:
                        _agregar_bloque(
                            estado, encabezado_ganador["documento"], encabezado_ganador["pagina"],
                            encabezado_ganador["franja"], resaltar=False, repetir_encabezado=False,
                        )

                _agregar_bloque(estado, documento_ganador, pagina_ganadora, franja, resaltar=resaltar_filas)

                polizas_encontradas.setdefault(clave, set()).add(Path(ruta_pdf_ganador).name)
                encontrados_por_nombre.add(clave)
            finally:
                documento_ganador.close()

    carpeta = Path(carpeta_salida)
    carpeta.mkdir(parents=True, exist_ok=True)
    archivos_por_cliente = {}
    for cliente, estado in estado_por_cliente.items():
        ruta = carpeta / f"{_nombre_archivo_seguro(cliente)}.pdf"
        estado["documento"].save(str(ruta))
        estado["documento"].close()
        archivos_por_cliente[cliente] = str(ruta)

    detalle_registros = []
    for clave, datos in registros_unicos.items():
        cliente = clave[1]
        polizas = sorted(polizas_encontradas.get(clave, set()))
        encontrado = bool(polizas)
        detalle_registros.append({
            "cedula": datos["cedula"],
            "nombre": datos["nombre"],
            "cliente": cliente,
            "polizas": polizas,
            "encontrado": encontrado,
            # "cedula": calzó por número; "nombre": rescatado en la segunda
            # pasada (más incierto -conviene revisarlo); None: no encontrado
            "encontrado_por": ("nombre" if clave in encontrados_por_nombre else "cedula") if encontrado else None,
        })

    no_encontrados = [
        {"cedula": d["cedula"], "cliente": d["cliente"], "nombre": d["nombre"]}
        for d in detalle_registros if not d["encontrado"]
    ]

    return {
        "archivos_por_cliente": archivos_por_cliente,
        "no_encontrados": no_encontrados,
        "errores_por_archivo": errores_por_archivo,
        "detalle_registros": detalle_registros,
    }


def generar_pdf_resumen(titulo: str, secciones: list[tuple[str, list[str]]]) -> bytes:
    """PDF simple de texto con listas por sección (ej. cédulas no encontradas,
    archivos con error), para meter dentro del zip en vez de mostrarlas en pantalla."""
    documento = fitz.open()
    pagina = documento.new_page()
    margen = 40
    alto_util = pagina.rect.height - margen

    y = margen
    pagina.insert_text((margen, y), titulo, fontsize=14)
    y += 26

    for encabezado, lineas in secciones:
        if not lineas:
            continue
        if y > alto_util - 20:
            pagina = documento.new_page()
            y = margen
        pagina.insert_text((margen, y), encabezado, fontsize=11)
        y += 18
        for linea in lineas:
            if y > alto_util:
                pagina = documento.new_page()
                y = margen
            pagina.insert_text((margen + 10, y), f"- {linea}", fontsize=9)
            y += 14
        y += 16

    datos = documento.tobytes()
    documento.close()
    return datos


_AZUL_VMA = "0A1F3D"
_RELLENO_ENCABEZADO = PatternFill("solid", fgColor=_AZUL_VMA)
_FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)
_BORDE_CELDA = Border(*(Side(style="thin", color="D7DBE3"),) * 4)


def _escribir_encabezado_excel(hoja, titulos: list[str]) -> None:
    hoja.append(titulos)
    for celda in hoja[1]:
        celda.fill = _RELLENO_ENCABEZADO
        celda.font = _FUENTE_ENCABEZADO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autoajustar_columnas_excel(hoja) -> None:
    for columna in hoja.columns:
        ancho = max((len(str(celda.value)) for celda in columna if celda.value is not None), default=8)
        hoja.column_dimensions[columna[0].column_letter].width = min(ancho + 4, 60)


def _bordear_filas_excel(hoja) -> None:
    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
        for celda in fila:
            celda.border = _BORDE_CELDA


def generar_excel_resumen(detalle_registros: list[dict]) -> bytes:
    """Arma el Excel de facturación (dos pestañas) que acompaña al zip del
    modo cliente, para que RRHH/Facturación no tenga que abrir los PDFs uno
    por uno a contar oficiales a mano.

    'detalle_registros' es la lista que devuelve
    resaltar_por_cedula_y_exportar_por_cliente bajo la llave del mismo
    nombre: un dict por cada (cédula, cliente) único del Excel, con
    'cedula', 'nombre', 'cliente', 'polizas' (archivos donde se encontró),
    'encontrado' y 'encontrado_por' ("cedula", "nombre" o None -"nombre"
    significa que se rescató en la segunda pasada y conviene revisarlo).

    Pestaña 1 ('Resumen por Cliente'): una fila por cliente con el total
    de oficiales listos para cobrar, cuántos faltan y en qué pólizas
    aparecieron -para copiar directo al sistema de facturación.
    Pestaña 2 ('Detalle de Oficiales'): una fila por oficial, para
    conciliar reclamos puntuales ("¿por qué a este cliente le falta un
    guarda?")."""
    libro = openpyxl.Workbook()

    hoja_resumen = libro.active
    hoja_resumen.title = "Resumen por Cliente"
    _escribir_encabezado_excel(hoja_resumen, [
        "Cliente", "Pólizas Involucradas", "Oficiales Encontrados", "Cédulas Faltantes", "Estado",
    ])

    por_cliente: dict[str, dict] = {}
    for d in detalle_registros:
        entrada = por_cliente.setdefault(d["cliente"], {"polizas": set(), "encontrados": 0, "faltantes": 0})
        if d["encontrado"]:
            entrada["encontrados"] += 1
            entrada["polizas"].update(d["polizas"])
        else:
            entrada["faltantes"] += 1

    total_encontrados = 0
    total_faltantes = 0
    for cliente in sorted(por_cliente):
        datos = por_cliente[cliente]
        estado = "🟢 Completo" if datos["faltantes"] == 0 else f"🟡 Pendiente ({datos['faltantes']})"
        hoja_resumen.append([
            cliente,
            ", ".join(sorted(datos["polizas"])) or "—",
            datos["encontrados"],
            datos["faltantes"],
            estado,
        ])
        total_encontrados += datos["encontrados"]
        total_faltantes += datos["faltantes"]

    hoja_resumen.append(["TOTAL GENERAL", "—", total_encontrados, total_faltantes, ""])
    for celda in hoja_resumen[hoja_resumen.max_row]:
        celda.font = Font(bold=True)
    _bordear_filas_excel(hoja_resumen)
    _autoajustar_columnas_excel(hoja_resumen)

    hoja_detalle = libro.create_sheet("Detalle de Oficiales")
    _escribir_encabezado_excel(hoja_detalle, [
        "Cédula", "Nombre Completo", "Cliente Asignado", "Póliza / Archivo de Origen", "¿Aparece en Planilla?",
    ])
    for d in sorted(detalle_registros, key=lambda d: (d["cliente"], d["cedula"])):
        if d.get("encontrado_por") == "nombre":
            estado_fila = "✅ Sí (por nombre -revisar)"
        elif d["encontrado"]:
            estado_fila = "✅ Sí"
        else:
            estado_fila = "❌ No encontrado"
        hoja_detalle.append([
            d["cedula"],
            d["nombre"] or "—",
            d["cliente"],
            ", ".join(d["polizas"]) if d["polizas"] else "(Ninguno)",
            estado_fila,
        ])
    _bordear_filas_excel(hoja_detalle)
    _autoajustar_columnas_excel(hoja_detalle)

    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def procesar_carpeta(ruta_carpeta_entrada: str, nombres: list[str], ruta_carpeta_salida: str) -> list[ResultadoResaltado]:
    """Aplica resaltar_nombres_en_pdf a todos los PDFs de una carpeta."""
    carpeta_entrada = Path(ruta_carpeta_entrada)
    carpeta_salida = Path(ruta_carpeta_salida)
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    resultados = []
    for archivo_pdf in sorted(carpeta_entrada.glob("*.pdf")):
        ruta_salida = carpeta_salida / f"{archivo_pdf.stem}_resaltado.pdf"
        resultado = resaltar_nombres_en_pdf(str(archivo_pdf), nombres, str(ruta_salida))
        resultados.append(resultado)

    return resultados
