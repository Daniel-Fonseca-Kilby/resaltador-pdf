"""Pruebas de integración de las rutas HTTP de api/index.py (Flask
test_client, sin arrancar un servidor real). Usa los mismos PDF/registros
inventados de conftest.py -nada de datos reales de empleados."""
import base64
import io
import json
import os
import shutil
import tempfile
import time
import zipfile
from pathlib import Path

import openpyxl
import pytest

from api.index import _limpiar_temporales_antiguos, app


@pytest.fixture
def cliente_flask():
    return app.test_client()


def _abrir_pdf(ruta):
    with open(ruta, "rb") as f:
        return io.BytesIO(f.read())


def _crear_excel_cliente(registros):
    """Excel con columnas Identificación/Cliente/Nombre, como lo subiría
    un usuario para el Modo Cliente."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["Identificacion", "Cliente", "Nombre"])
    for registro in registros:
        hoja.append([registro["cedula"], registro["cliente"], registro["nombre"]])
    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer


def test_index_devuelve_200(cliente_flask):
    respuesta = cliente_flask.get("/")
    assert respuesta.status_code == 200


def test_procesar_sin_pdfs_devuelve_400(cliente_flask):
    respuesta = cliente_flask.post("/api/procesar", data={"nombres": "Juan Perez"})

    assert respuesta.status_code == 400
    assert respuesta.is_json
    assert "PDF" in respuesta.get_json()["error"]


def test_procesar_excel_no_valido_devuelve_400(cliente_flask, ruta_pdf_ejemplo):
    datos = {
        "nombres": "",
        "excel": (io.BytesIO(b"esto no es un excel"), "lista.txt"),
        "pdfs": (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
    }

    respuesta = cliente_flask.post("/api/procesar", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 400
    assert "Excel" in respuesta.get_json()["error"]


def test_procesar_modo_simple_devuelve_zip(cliente_flask, ruta_pdf_ejemplo):
    datos = {
        "nombres": "Juan Perez",
        "pdfs": (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
    }

    respuesta = cliente_flask.post("/api/procesar", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 200
    assert respuesta.headers["Content-Type"] == "application/zip"
    assert respuesta.headers["X-Modo"] == "simple"
    assert respuesta.headers["X-Total-Archivos"] == "1"


def test_procesar_modo_cliente_devuelve_zip(cliente_flask, ruta_pdf_ejemplo, registros_ejemplo):
    datos = {
        "excel": (_crear_excel_cliente(registros_ejemplo), "planilla.xlsx"),
        "pdfs": (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
        "formato": "mnk",
    }

    respuesta = cliente_flask.post("/api/procesar", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 200
    assert respuesta.headers["Content-Type"] == "application/zip"
    assert respuesta.headers["X-Modo"] == "cliente"
    assert respuesta.headers["X-Total-Clientes"] == "2"


def test_procesar_modo_cliente_manda_no_encontrados_en_cabecera(cliente_flask, ruta_pdf_ejemplo, registros_ejemplo):
    # registros_ejemplo trae una cédula (999999999) que no está en el PDF
    datos = {
        "excel": (_crear_excel_cliente(registros_ejemplo), "planilla.xlsx"),
        "pdfs": (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
        "formato": "mnk",
    }

    respuesta = cliente_flask.post("/api/procesar", data=datos, content_type="multipart/form-data")

    assert respuesta.headers["X-Total-No-Encontrados"] == "1"
    lista = json.loads(base64.b64decode(respuesta.headers["X-No-Encontrados-B64"]).decode("utf-8"))
    assert len(lista) == 1
    assert "999999999" in lista[0]


def test_procesar_multiples_pdfs_mismo_nombre(cliente_flask, ruta_pdf_ejemplo):
    # Es común descargar "Planilla.pdf" de dos portales distintos (ej. CCSS
    # de San José y de Heredia) con el mismo nombre de archivo -ninguno de
    # los dos se debe perder ni pisar al otro.
    datos = {
        "nombres": "Juan Perez",
        "pdfs": [
            (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
            (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
        ],
    }

    respuesta = cliente_flask.post("/api/procesar", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 200
    assert respuesta.headers["X-Total-Archivos"] == "2"

    with zipfile.ZipFile(io.BytesIO(respuesta.data)) as zf:
        nombres_pdf = [n for n in zf.namelist() if n.lower().endswith(".pdf")]

    assert len(nombres_pdf) == 2
    assert len(set(nombres_pdf)) == 2  # nombres distintos, ninguno se pisó


def test_procesar_modo_cliente_con_csv_delimitado_por_comas(cliente_flask, ruta_pdf_ejemplo):
    contenido_csv = (
        "Identificacion,Cliente,Nombre\n"
        "111111111,Cliente Prueba Uno,Juan Perez Mora\n"
        "222222222,Cliente Prueba Uno,Maria Jose Solano\n"
    ).encode("utf-8")

    datos = {
        "excel": (io.BytesIO(contenido_csv), "planillas.csv"),
        "pdfs": (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
    }
    respuesta = cliente_flask.post("/api/procesar", data=datos, content_type="multipart/form-data")
    assert respuesta.status_code == 200
    assert respuesta.headers["X-Modo"] == "cliente"


def test_procesar_modo_cliente_con_csv_punto_y_coma_latin1(cliente_flask, ruta_pdf_ejemplo):
    # CSV con punto y coma y caracteres con tilde/ñ en Latin-1 (exportación típica de ERP)
    contenido_csv = (
        "Cédula;Empresa;Colaborador\n"
        "111111111;Cliente Prueba Uno;Juan Pérez\n"
    ).encode("latin-1")

    datos = {
        "excel": (io.BytesIO(contenido_csv), "exportacion_erp.csv"),
        "pdfs": (_abrir_pdf(ruta_pdf_ejemplo), "planilla.pdf"),
    }
    respuesta = cliente_flask.post("/api/procesar", data=datos, content_type="multipart/form-data")
    assert respuesta.status_code == 200
    assert respuesta.headers["X-Modo"] == "cliente"


def test_detectar_modo_excel_con_csv(cliente_flask):
    contenido_csv = (
        "Identificacion,Cliente,Nombre\n"
        "111111111,Cliente Prueba Uno,Juan Perez Mora\n"
    ).encode("utf-8")

    datos = {"excel": (io.BytesIO(contenido_csv), "planillas.csv")}
    respuesta = cliente_flask.post("/api/detectar-modo-excel", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 200
    cuerpo = respuesta.get_json()
    assert cuerpo["modo"] == "cliente"
    assert cuerpo["total_registros"] == 1


def test_detectar_modo_excel_reconoce_columnas_de_cliente(cliente_flask, registros_ejemplo):
    datos = {"excel": (_crear_excel_cliente(registros_ejemplo), "planilla.xlsx")}

    respuesta = cliente_flask.post("/api/detectar-modo-excel", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 200
    cuerpo = respuesta.get_json()
    assert cuerpo["modo"] == "cliente"
    assert cuerpo["total_registros"] == 4


def test_detectar_modo_excel_una_sola_columna_es_modo_simple(cliente_flask):
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["Nombre"])
    hoja.append(["Juan Perez"])
    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    datos = {"excel": (buffer, "nombres.xlsx")}
    respuesta = cliente_flask.post("/api/detectar-modo-excel", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 200
    assert respuesta.get_json()["modo"] == "simple"


def test_detectar_modo_excel_columnas_no_reconocidas_da_error(cliente_flask):
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["Fecha", "Comentario"])
    hoja.append(["2026-01-01", "Nota"])
    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    datos = {"excel": (buffer, "otro.xlsx")}
    respuesta = cliente_flask.post("/api/detectar-modo-excel", data=datos, content_type="multipart/form-data")

    assert respuesta.status_code == 400
    assert "no contiene una columna" in respuesta.get_json()["error"]


def test_limpiar_temporales_antiguos_borra_viejos_y_conserva_nuevos():
    temp_dir = Path(tempfile.gettempdir())

    # 1. Crear carpeta vieja simulada (antigüedad de 2 horas)
    carpeta_vieja = temp_dir / "resaltado_cliente_test_antigua"
    carpeta_vieja.mkdir(parents=True, exist_ok=True)
    hace_dos_horas = time.time() - 7200
    os.utime(str(carpeta_vieja), (hace_dos_horas, hace_dos_horas))

    # 2. Crear carpeta reciente (antigüedad de 5 segundos)
    carpeta_nueva = temp_dir / "resaltado_cliente_test_reciente"
    carpeta_nueva.mkdir(parents=True, exist_ok=True)

    try:
        # Ejecutar limpieza con límite de 1 hora (3600s)
        _limpiar_temporales_antiguos(segundos_vida=3600)

        # Verificar: la vieja debió borrarse y la nueva conservarse
        assert not carpeta_vieja.exists()
        assert carpeta_nueva.exists()
    finally:
        # Limpieza manual del fixture
        if carpeta_vieja.exists():
            shutil.rmtree(carpeta_vieja, ignore_errors=True)
        if carpeta_nueva.exists():
            shutil.rmtree(carpeta_nueva, ignore_errors=True)
