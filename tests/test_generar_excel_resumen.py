"""Pruebas de generar_excel_resumen (el Excel de facturación de dos
pestañas que va dentro del zip del modo cliente)."""
from io import BytesIO

import openpyxl

from resaltado_pdf import generar_excel_resumen


def _detalle_ejemplo():
    return [
        {
            "cedula": "111111111", "nombre": "Juan Perez", "cliente": "Walmart",
            "polizas": ["poliza_a.pdf"], "encontrado": True,
        },
        {
            "cedula": "222222222", "nombre": "Maria Solano", "cliente": "Walmart",
            "polizas": ["poliza_a.pdf"], "encontrado": True,
        },
        {
            "cedula": "999999999", "nombre": "Carlos Zuniga", "cliente": "Banco BAC",
            "polizas": [], "encontrado": False,
        },
    ]


def test_genera_excel_con_las_dos_pestanas_esperadas():
    datos = generar_excel_resumen(_detalle_ejemplo())
    assert isinstance(datos, bytes)

    libro = openpyxl.load_workbook(BytesIO(datos))
    assert libro.sheetnames == ["Resumen por Cliente", "Detalle de Oficiales"]


def test_resumen_por_cliente_cuenta_encontrados_y_faltantes():
    datos = generar_excel_resumen(_detalle_ejemplo())
    hoja = openpyxl.load_workbook(BytesIO(datos))["Resumen por Cliente"]

    filas = list(hoja.iter_rows(values_only=True))
    filas_por_cliente = {fila[0]: fila for fila in filas[1:-1]}  # sin encabezado ni TOTAL GENERAL

    assert filas_por_cliente["Walmart"][2] == 2  # oficiales encontrados
    assert filas_por_cliente["Walmart"][3] == 0  # cédulas faltantes
    assert filas_por_cliente["Banco BAC"][2] == 0
    assert filas_por_cliente["Banco BAC"][3] == 1

    fila_total = filas[-1]
    assert fila_total[0] == "TOTAL GENERAL"
    assert fila_total[2] == 2
    assert fila_total[3] == 1


def test_detalle_de_oficiales_marca_encontrado_y_no_encontrado():
    datos = generar_excel_resumen(_detalle_ejemplo())
    hoja = openpyxl.load_workbook(BytesIO(datos))["Detalle de Oficiales"]

    filas = list(hoja.iter_rows(values_only=True))[1:]  # sin encabezado
    por_cedula = {fila[0]: fila for fila in filas}

    assert "Sí" in por_cedula["111111111"][4]
    assert por_cedula["111111111"][3] == "poliza_a.pdf"

    assert "No encontrado" in por_cedula["999999999"][4]
    assert por_cedula["999999999"][3] == "(Ninguno)"


def test_generar_excel_resumen_con_lista_vacia_no_falla():
    datos = generar_excel_resumen([])
    libro = openpyxl.load_workbook(BytesIO(datos))
    assert libro.sheetnames == ["Resumen por Cliente", "Detalle de Oficiales"]
